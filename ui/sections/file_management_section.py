"""
قسم إدارة الملفات - File Management Section
يتضمن منطقة رفع ملفات Excel ومعاينة معلومات الملف
"""
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                               QPushButton, QFileDialog, QFrame, 
                               QGraphicsBlurEffect, QGraphicsDropShadowEffect)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont, QColor, QDragEnterEvent, QDropEvent
import os


class FileManagementSection(QWidget):
    """قسم إدارة الملفات مع منطقة السحب والإفلات ومعاينة الملف"""
    
    # Signal to emit when a file is selected
    file_selected = Signal(str)  # file_path
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_file_path = None
        self.file_info = {
            'name': '',
            'rows': 0,
            'columns': 0
        }
        self._setup_ui()
        self._apply_styles()
        
    def _setup_ui(self):
        """إعداد واجهة المستخدم"""
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Create background blur layer for glass effect
        self.blur_background = QFrame()
        self.blur_background.setObjectName("blurBackground")
        self.blur_background.setMinimumHeight(450)
        
        # Apply blur effect to background
        blur_effect = QGraphicsBlurEffect()
        blur_effect.setBlurRadius(15)
        blur_effect.setBlurHints(QGraphicsBlurEffect.QualityHint)
        self.blur_background.setGraphicsEffect(blur_effect)
        
        # Card container with vertical layout
        self.card = QFrame()
        self.card.setObjectName("fileManagementCard")
        card_main_layout = QVBoxLayout(self.card)
        card_main_layout.setContentsMargins(30, 25, 30, 30)
        card_main_layout.setSpacing(20)
        
        # Main Title - File Management (inside card now)
        title_layout = QHBoxLayout()
        title_layout.setSpacing(10)
        
        # Excel file icon (SVG implementation)
        icon_label = QLabel()
        icon_label.setAlignment(Qt.AlignCenter)
        
        # Create Excel icon SVG
        excel_svg = """
        <svg width="24" height="24" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
            <!-- Excel file background -->
            <rect x="3" y="2" width="14" height="20" rx="1" fill="#217346"/>
            <!-- Corner fold -->
            <path d="M 17,2 L 17,7 L 22,7 Z" fill="#185C37"/>
            <path d="M 17,2 L 22,7 L 17,7 Z" fill="#21A366"/>
            <!-- White X -->
            <line x1="6" y1="10" x2="11" y2="16" stroke="white" stroke-width="1.5" stroke-linecap="round"/>
            <line x1="11" y1="10" x2="6" y2="16" stroke="white" stroke-width="1.5" stroke-linecap="round"/>
            <!-- Sheet lines -->
            <line x1="6" y1="18" x2="14" y2="18" stroke="white" stroke-width="0.5" opacity="0.5"/>
            <line x1="6" y1="20" x2="14" y2="20" stroke="white" stroke-width="0.5" opacity="0.5"/>
        </svg>
        """
        
        from PySide6.QtSvg import QSvgRenderer
        from PySide6.QtGui import QPixmap, QPainter
        
        renderer = QSvgRenderer(excel_svg.encode())
        pixmap = QPixmap(24, 24)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        renderer.render(painter)
        painter.end()
        
        icon_label.setPixmap(pixmap)
        title_layout.addWidget(icon_label)
        
        # Title
        title = QLabel("File Management")
        title.setObjectName("mainTitle")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_layout.addWidget(title)
        title_layout.addStretch()
        
        card_main_layout.addLayout(title_layout)
        
        # Horizontal layout for the two boxes
        boxes_layout = QHBoxLayout()
        boxes_layout.setSpacing(40)
        
        # Left box - Upload area only (no title)
        left_box = self._create_upload_box()
        boxes_layout.addWidget(left_box, 1)
        
        # Right box - File Preview
        right_box = self._create_preview_section()
        boxes_layout.addWidget(right_box, 1)
        
        card_main_layout.addLayout(boxes_layout)
        
        main_layout.addWidget(self.card)
        
    def _create_upload_box(self):
        """إنشاء مربع رفع الملفات (بدون عنوان)"""
        # Upload area
        self.upload_area = QFrame()
        self.upload_area.setObjectName("uploadArea")
        self.upload_area.setMinimumHeight(220)
        self.upload_area.setAcceptDrops(True)
        
        # Override drag and drop events
        self.upload_area.dragEnterEvent = self._drag_enter_event
        self.upload_area.dragLeaveEvent = self._drag_leave_event
        self.upload_area.dropEvent = self._drop_event
        
        upload_layout = QVBoxLayout(self.upload_area)
        upload_layout.setAlignment(Qt.AlignCenter)
        upload_layout.setSpacing(10)
        
        # Cloud icon with upload arrow inside (SVG implementation)
        cloud_icon = QLabel()
        cloud_icon.setObjectName("cloudIcon")
        cloud_icon.setAlignment(Qt.AlignCenter)
        
        # Create SVG icon
        svg_data = """
        <svg width="100" height="80" viewBox="0 0 120 100" xmlns="http://www.w3.org/2000/svg">
            <!-- Cloud shape -->
            <path d="M 85,45 Q 85,35 75,30 Q 75,20 60,20 Q 45,20 40,30 Q 30,30 25,40 Q 20,45 25,55 Q 25,65 35,65 L 85,65 Q 95,65 95,55 Q 95,45 85,45 Z" 
                  fill="#6B4EEB" stroke="none"/>
            <!-- Upload arrow inside cloud -->
            <g transform="translate(60, 50)">
                <!-- Arrow shaft -->
                <rect x="-4" y="-10" width="8" height="25" fill="white"/>
                <!-- Arrow head -->
                <polygon points="0,-15 -10,-5 10,-5" fill="white"/>
            </g>
        </svg>
        """
        
        from PySide6.QtSvg import QSvgRenderer
        from PySide6.QtGui import QPixmap, QPainter
        
        # Render SVG to pixmap
        renderer = QSvgRenderer(svg_data.encode())
        pixmap = QPixmap(100, 80)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        renderer.render(painter)
        painter.end()
        
        cloud_icon.setPixmap(pixmap)
        upload_layout.addWidget(cloud_icon)
        
        # Main instruction
        main_text = QLabel("Import Excel File")
        main_text.setObjectName("uploadMainText")
        main_text.setFont(QFont("Segoe UI", 12, QFont.DemiBold))
        main_text.setAlignment(Qt.AlignCenter)
        upload_layout.addWidget(main_text)
        
        # Secondary instruction
        secondary_text = QLabel("Drag and drop or click to select your Excel file")
        secondary_text.setObjectName("uploadSecondaryText")
        secondary_text.setFont(QFont("Segoe UI", 9))
        secondary_text.setAlignment(Qt.AlignCenter)
        secondary_text.setWordWrap(True)
        upload_layout.addWidget(secondary_text)
        
        # Choose File button
        self.choose_button = QPushButton("Choose File")
        self.choose_button.setObjectName("chooseFileButton")
        self.choose_button.setFont(QFont("Segoe UI", 8, QFont.Bold))
        self.choose_button.setMinimumHeight(35)
        self.choose_button.setMinimumWidth(120)
        self.choose_button.setCursor(Qt.PointingHandCursor)
        self.choose_button.clicked.connect(self._choose_file)
        upload_layout.addWidget(self.choose_button, alignment=Qt.AlignCenter)
        
        return self.upload_area
        
    def _create_preview_section(self):
        """إنشاء قسم معاينة الملف"""
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)
        
        # File information container with white background
        self.info_container = QFrame()
        self.info_container.setObjectName("previewContainer")
        self.info_container.setMinimumHeight(220)
        info_layout = QVBoxLayout(self.info_container)
        info_layout.setContentsMargins(20, 15, 20, 15)
        info_layout.setSpacing(10)
        
        # Header with icon and title (inside the container - as subtitle)
        header_layout = QHBoxLayout()
        header_layout.setSpacing(8)
        
        # Eye icon (smaller)
        icon_label = QLabel("👁️")
        icon_label.setFont(QFont("Segoe UI", 14))
        header_layout.addWidget(icon_label)
        
        # Title (smaller - subtitle style)
        title = QLabel("File Preview")
        title.setObjectName("previewSubtitle")
        title.setFont(QFont("Segoe UI", 11, QFont.DemiBold))
        header_layout.addWidget(title)
        header_layout.addStretch()
        
        info_layout.addLayout(header_layout)
        
        # Separator line (thinner)
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setObjectName("previewSeparator")
        info_layout.addWidget(separator)
        
        # Add stretch to push content to center
        info_layout.addStretch()
        
        # File Name
        self.file_name_layout = self._create_info_row("File Name:", "No file selected")
        info_layout.addLayout(self.file_name_layout)
        
        # Rows
        self.rows_layout = self._create_info_row("Rows:", "0")
        info_layout.addLayout(self.rows_layout)
        
        # Columns
        self.columns_layout = self._create_info_row("Columns:", "0")
        info_layout.addLayout(self.columns_layout)
        
        # Add stretch to push content to center
        info_layout.addStretch()
        
        layout.addWidget(self.info_container)
        
        return container
        
    def _create_info_row(self, label_text, value_text):
        """إنشاء صف معلومات (تسمية + قيمة)"""
        row_layout = QHBoxLayout()
        row_layout.setSpacing(10)
        
        # Label
        label = QLabel(label_text)
        label.setObjectName("infoLabel")
        label.setFont(QFont("Segoe UI", 11))
        row_layout.addWidget(label)
        
        # Value
        value = QLabel(value_text)
        value.setObjectName("infoValue")
        value.setFont(QFont("Segoe UI", 11, QFont.Bold))
        row_layout.addWidget(value)
        
        row_layout.addStretch()
        
        return row_layout
        
    def _choose_file(self):
        """فتح نافذة اختيار الملف"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*.*)"
        )
        
        if file_path:
            self._load_file(file_path)
            
    def _load_file(self, file_path):
        """تحميل الملف وتحديث المعلومات"""
        try:
            self.current_file_path = file_path
            file_name = os.path.basename(file_path)
            
            # Try to read file info using openpyxl or similar
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                rows = ws.max_row
                columns = ws.max_column
                wb.close()
            except:
                # If openpyxl is not available, show basic info
                rows = 0
                columns = 0
            
            # Update file info
            self.file_info = {
                'name': file_name,
                'rows': rows,
                'columns': columns
            }
            
            # Update preview labels
            self._update_preview()
            
            # Emit signal
            self.file_selected.emit(file_path)
            
        except Exception as e:
            print(f"Error loading file: {e}")
            
    def _update_preview(self):
        """تحديث معلومات المعاينة"""
        # Update file name
        file_name_value = self.file_name_layout.itemAt(1).widget()
        file_name_value.setText(self.file_info['name'])
        
        # Update rows
        rows_value = self.rows_layout.itemAt(1).widget()
        rows_value.setText(f"{self.file_info['rows']:,}")
        
        # Update columns
        columns_value = self.columns_layout.itemAt(1).widget()
        columns_value.setText(str(self.file_info['columns']))
        
    def _drag_enter_event(self, event: QDragEnterEvent):
        """عند سحب ملف فوق منطقة الرفع"""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.upload_area.setProperty("dragActive", True)
            self.upload_area.style().unpolish(self.upload_area)
            self.upload_area.style().polish(self.upload_area)
            
    def _drag_leave_event(self, event):
        """عند مغادرة الملف لمنطقة الرفع"""
        self.upload_area.setProperty("dragActive", False)
        self.upload_area.style().unpolish(self.upload_area)
        self.upload_area.style().polish(self.upload_area)
        
    def _drop_event(self, event: QDropEvent):
        """عند إفلات ملف في منطقة الرفع"""
        self.upload_area.setProperty("dragActive", False)
        self.upload_area.style().unpolish(self.upload_area)
        self.upload_area.style().polish(self.upload_area)
        
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if file_path.endswith(('.xlsx', '.xls')):
                    self._load_file(file_path)
                    event.acceptProposedAction()
                    
    def _apply_styles(self):
        """تطبيق الأنماط CSS مع تأثير Glassmorphism"""
        # Add soft shadow effect for the card
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(35)
        shadow.setXOffset(0)
        shadow.setYOffset(15)
        shadow.setColor(QColor(107, 78, 235, 60))  # Purple shadow matching theme
        self.card.setGraphicsEffect(shadow)
        
        self.setStyleSheet("""
            /* Blur background layer */
            #blurBackground {
                background-color: rgba(255, 255, 255, 0.1);
                border-radius: 18px;
            }
            
            /* Card container with Glassmorphism effect - Softer borders */
            #fileManagementCard {
                background-color: rgba(255, 255, 255, 0.25);
                border-radius: 18px;
                border: 0.5px solid rgba(255, 255, 255, 0.4);
            }
            
            /* Main Title (outside card) */
            #mainTitle {
                color: #1a1a1a;
                font-weight: 700;
            }
            
            /* Section titles */
            #sectionTitle {
                color: #1a1a1a;
                font-weight: 600;
            }
            
            /* Preview subtitle (smaller) */
            #previewSubtitle {
                color: #4a4a4a;
                font-weight: 600;
            }
            
            /* Preview Container with white background */
            #previewContainer {
                background-color: rgba(255, 255, 255, 0.7);
                border-radius: 12px;
                border: 0.5px solid rgba(220, 220, 220, 0.3);
            }
            
            /* Separator line in preview */
            #previewSeparator {
                background-color: rgba(200, 200, 200, 0.3);
                max-height: 1px;
                margin: 5px 0px;
            }
            
            /* Upload area */
            #uploadArea {
                background-color: rgba(248, 249, 250, 0.5);
                border: 2px dashed rgba(204, 204, 204, 0.8);
                border-radius: 12px;
            }
            
            #uploadArea[dragActive="true"] {
                background-color: rgba(232, 234, 255, 0.75);
                border-color: #6B4EEB;
                border-width: 2px;
                border-style: solid;
            }
            
            /* Cloud icon */
            #cloudIcon {
                color: #6B4EEB;
            }
            
            /* Upload text */
            #uploadMainText {
                color: #1a1a1a;
                font-weight: 600;
            }
            
            #uploadSecondaryText {
                color: #4a4a4a;
            }
            
            /* Choose file button with glass effect */
            #chooseFileButton {
                background-color: rgba(107, 78, 235, 0.9);
                color: #FFFFFF;
                border: 1px solid rgba(255, 255, 255, 0.3);
                border-radius: 10px;
                padding: 12px 30px;
                font-weight: 600;
            }
            
            #chooseFileButton:hover {
                background-color: rgba(90, 61, 217, 0.95);
                border: 1px solid rgba(255, 255, 255, 0.5);
            }
            
            #chooseFileButton:pressed {
                background-color: rgba(75, 46, 199, 1);
                border: 1px solid rgba(255, 255, 255, 0.2);
            }
            
            /* Info labels */
            #infoLabel {
                color: #2a2a2a;
            }
            
            #infoValue {
                color: #1a1a1a;
                font-weight: 600;
            }
        """)
        
    def get_file_path(self):
        """الحصول على مسار الملف الحالي"""
        return self.current_file_path
        
    def get_file_info(self):
        """الحصول على معلومات الملف"""
        return self.file_info
